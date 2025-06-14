from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from .models import Customer, Product, Stock, Supplier, Category, Sale, SaleProduct, SalePayment, Purchase, PurchaseProduct, Boleto, Manufacturer
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
import requests
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum, Count, Min
from datetime import datetime
from django.utils.timezone import now, timedelta, make_aware, localtime, is_naive, get_current_timezone
from datetime import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.contrib import messages

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import reportlab
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph
from textwrap import wrap
from io import BytesIO
from reportlab.lib.styles import getSampleStyleSheet


from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from app_stock_manager.decorators import admin_required  # Importa o decorador personalizado

from django.db.models import Q



# Função para verificar se o usuário é admin
def is_admin(user):
    return user.is_staff  # Somente usuários com is_staff=True podem acessar

def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("home")  # Redireciona para a página inicial
        else:
            messages.error(request, "Nome de usuário ou senha incorretos.")
    return render(request, "pages/login.html")

class CustomPasswordChangeView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'pages/change_password.html'
    success_url = reverse_lazy('home')
    success_message = "Sua senha foi alterada com sucesso!"

    def form_invalid(self, form):
        # Adiciona mensagens de erro específicas do formulário
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, error)
        return super().form_invalid(form)

@login_required
def home(request):
    return render(request, 'pages/home.html')

@login_required
def settings(request):
    return render(request, 'pages/settings.html')

@login_required
@admin_required
def new_user(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        is_admin = request.POST.get("is_admin") == "on"  # Verifica se a checkbox foi marcada

        # Verifica se as senhas coincidem
        if password != confirm_password:
            messages.error(request, "As senhas não coincidem.")
            return redirect("new_user")

        # Verifica se o nome de usuário já existe
        if User.objects.filter(username=username).exists():
            messages.error(request, "Nome de usuário já existe.")
            return redirect("new_user")

        # Criar usuário
        user = User.objects.create_user(username=username, password=password)
        if is_admin:
            user.is_staff = True  # Define como administrador
            user.is_superuser = True  # Torna também superusuário, se necessário
        user.save()

        messages.success(request, "Usuário criado com sucesso!")
        return redirect("new_user")
    return render(request, 'pages/new_user.html')

@login_required
def view_customers(request):
    return render(request, 'pages/list_customers.html')

@login_required
def new_customer(request):
    return render(request, 'pages/new_customer.html')

@login_required
def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == "POST":
        customer.name = request.POST.get("name")
        customer.contact_name = request.POST.get("contact_name")
        customer.phone = request.POST.get("phone")
        customer.email = request.POST.get("email")
        customer.cep = request.POST.get("cep")
        customer.street = request.POST.get("street")
        customer.address_number = request.POST.get("address_number")
        customer.district = request.POST.get("district")
        customer.state = request.POST.get("state")
        customer.city = request.POST.get("city")
        customer.cpf_or_cnpj = request.POST.get("cpf_or_cnpj")
        customer.is_cnpj = request.POST.get("is_cnpj") == "true"
        customer.observations = request.POST.get("observations")

        customer.save()
        messages.success(request, "Cliente atualizado com sucesso!")
        return redirect("view_customers")

    return render(request, "pages/edit_customer.html", {"customer": customer})

@login_required
def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    customer.delete()
    return redirect("view_customers")  # Redireciona para a lista de clientes

@login_required
def delete_manufacturer(request, manufacturer_id):
    manufacturer = get_object_or_404(Manufacturer, id=manufacturer_id)
    manufacturer.delete()
    return redirect("view_manufactures")  # Redireciona para a lista de clientes

@login_required
def delete_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    category.delete()
    return redirect("view_categorys") 

@login_required
def view_suppliers(request):
    return render(request, 'pages/list_suppliers.html')

@login_required
def new_supplier(request):
    return render(request, 'pages/new_supplier.html')

@login_required
def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)

    if request.method == "POST":
        supplier.name = request.POST.get("name")
        supplier.cnpj = request.POST.get("cnpj")
        supplier.contact_name = request.POST.get("contact_name")
        supplier.phone = request.POST.get("phone")
        supplier.email = request.POST.get("email")
        supplier.cep = request.POST.get("cep")
        supplier.street = request.POST.get("street")
        supplier.address_number = request.POST.get("address_number")
        supplier.district = request.POST.get("district")
        supplier.state = request.POST.get("state")
        supplier.city = request.POST.get("city")
        supplier.observations = request.POST.get("observations")

        supplier.save()
        messages.success(request, "Fornecedor atualizado com sucesso!")
        return redirect("view_suppliers")

    return render(request, "pages/edit_supplier.html", {"supplier": supplier})


    return render(request, "pages/edit_supplier.html", {"supplier": supplier})

@login_required
def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    supplier.delete()
    return redirect("view_suppliers")  # Redireciona para a lista de clientes

@login_required
def view_products(request):
    return render(request, 'pages/list_products.html')

@login_required
def new_product(request):
    return render(request, 'pages/new_product.html')

@csrf_exempt
@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'pages/edit_product.html',{"product": product})

@login_required
def view_stocks(request):
    return render(request, 'pages/list_stock.html')

@login_required
def view_sales(request):
    return render(request, 'pages/list_sales.html')

@login_required
def new_sale(request):
    return render(request, 'pages/new_sale.html')

@login_required
def view_purchases(request):
    return render(request, 'pages/list_purchases.html')

@login_required
def new_purchase(request):
    return render(request, 'pages/new_purchase.html')

@login_required
def view_manufactures(request):
    return render(request, 'pages/list_manufacturer.html')

@login_required
def new_manufacturer(request):
    return render(request, 'pages/new_manufacturer.html')

@login_required
def view_categorys(request):
    return render(request, 'pages/list_category.html')

@login_required
def new_category(request):
    return render(request, 'pages/new_category.html')

# ======================= FUNCIONS ============================================
def user_logout(request):
    logout(request)
    return redirect("login")

@login_required
def add_category(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()

        if not name:
            return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)

        try:
            category = Categories.objects.create(
                name=name,
            )
            return JsonResponse({"message": "Categoria adicionado com sucesso!", "id": category.id})
        except Exception as e:
            return JsonResponse({"error": f"Erro ao salvar categoria: {str(e)}"}, status=500)
    else:
        return redirect('view_categorys')

@login_required
def add_customer(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        contact_name = request.POST.get("contact_name", "").strip()
        phone = request.POST.get("phone", "").strip()
        email = request.POST.get("email", "").strip()
        cep = request.POST.get("cep", "").strip()
        street = request.POST.get("street", "").strip()
        address_number = request.POST.get("address_number", "").strip()
        district = request.POST.get("district", "").strip()
        state = request.POST.get("state", "").strip()
        city = request.POST.get("city", "").strip()
        cpf_or_cnpj = request.POST.get("cpf_or_cnpj", "").strip()
        is_cnpj = request.POST.get("is_cnpj", "").lower() == "true"
        observations = request.POST.get("observations", "").strip()

        if not name:
            return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)

        try:
            customer = Customer.objects.create(
                name=name,
                contact_name=contact_name,
                phone=phone,
                email=email,
                cpf_or_cnpj=cpf_or_cnpj,
                is_cnpj=is_cnpj,
                street=street,
                address_number=address_number,
                district=district,
                city=city,
                state=state,
                cep=cep,
                observations=observations,
            )
            return redirect('view_customers')
        except Exception as e:
            return JsonResponse({"error": f"Erro ao salvar cliente: {str(e)}"}, status=500)
    else:
        return redirect('home')


@csrf_exempt
@login_required
def add_supplier(request):
    if request.method == "POST":
        try:
            name = request.POST.get("name", "").strip()
            cnpj = request.POST.get("cnpj", "").strip()
            contact_name = request.POST.get("contact_name", "").strip()
            phone = request.POST.get("phone", "").strip()
            email = request.POST.get("email", "").strip()
            street = request.POST.get("street", "").strip()
            address_number = request.POST.get("address_number", "").strip()
            district = request.POST.get("district", "").strip()
            city = request.POST.get("city", "").strip()
            state = request.POST.get("state", "").strip()
            cep = request.POST.get("cep", "").strip()
            observations = request.POST.get("observations", "").strip()

            if not name:
                return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)

            supplier = Supplier.objects.create(
                name=name,
                cnpj=cnpj,
                contact_name=contact_name,
                phone=phone,
                email=email,
                street=street,
                address_number=address_number,
                district=district,
                city=city,
                state=state,
                cep=cep,
                observations=observations,
            )

            return redirect('view_suppliers')

        except Exception as e:
            return JsonResponse({"error": f"Erro ao salvar fornecedor: {str(e)}"}, status=500)

    else:
        return redirect('home')

@csrf_exempt
@login_required
def add_product(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            name = data.get("name", "").strip()
            weight = data.get("weight", "").strip()
            unit_type = data.get("unit_type", "").strip()
            size = data.get("size", "").strip()
            quantity_per_package = data.get("quantity_per_package", "").strip()
            observations = data.get("observations", "").strip()

            manufacturer_id = data.get("manufacturer")
            category_id = data.get("category")

            # Validações
            if not name:
                return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)
            if not weight.isdigit() or int(weight) <= 0:
                return JsonResponse({"error": "O campo 'Gramas' deve ser um número positivo."}, status=400)
            if not size or "x" not in size:
                return JsonResponse({"error": "O campo 'Formato' deve ser no formato 'NxM'."}, status=400)
            if not quantity_per_package.isdigit() or int(quantity_per_package) <= 0:
                return JsonResponse({"error": "O campo 'Quantidade por Pacote' deve ser um número positivo."}, status=400)

            manufacturer = Manufacturer.objects.filter(id=manufacturer_id).first()
            category = Category.objects.filter(id=category_id).first()

            product = Product.objects.create(
                name=name,
                weight=int(weight),
                unit_type=unit_type,
                size=size,
                quantity_per_package=int(quantity_per_package),
                observations=observations,
                manufacturer=manufacturer,
                category=category,
            )

            Stock.objects.create(
                product=product,
                quantity=0,
                price=0.0,
            )

            return JsonResponse({"message": "Produto adicionado com sucesso!", "id": product.id})

        except json.JSONDecodeError as e:
            return JsonResponse({"error": f"Dados JSON inválidos: {str(e)}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Erro ao processar a solicitação: {str(e)}"}, status=500)

    return JsonResponse({"error": "Método não permitido."}, status=405)

@csrf_exempt
@login_required
def update_product(request, product_id):
    if request.method == "POST":
        try:
            product = get_object_or_404(Product, id=product_id)
            data = json.loads(request.body)

            product.name = data.get("name", product.name).strip()
            product.weight = int(data.get("weight", product.weight))
            product.unit_type = data.get("unit_type", product.unit_type)
            product.size = data.get("size", product.size)
            product.quantity_per_package = int(data.get("quantity_per_package", product.quantity_per_package))
            product.observations = data.get("observations", product.observations)

            manufacturer_id = data.get("manufacturer")
            category_id = data.get("category")

            product.manufacturer = Manufacturer.objects.filter(id=manufacturer_id).first() if manufacturer_id else None
            product.category = Category.objects.filter(id=category_id).first() if category_id else None

            product.save()

            return JsonResponse({"message": "Produto atualizado com sucesso!"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método não permitido"}, status=405)


@csrf_exempt
@login_required
def add_sale(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método não permitido"}, status=405)

    try:
        data = json.loads(request.body)

        customer_id = data.get("customer")

        # Validação explícita do cliente
        if not customer_id:
            return JsonResponse({"success": False, "error": "Selecione um Cliente"}, status=400)

        try:
            customer_id_int = int(customer_id)
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "error": "Selecione um Cliente válido"}, status=400)

        customer = get_object_or_404(Customer, id=customer_id_int)

        # restante do seu código abaixo, sem alterações
        # ...
        # (copie seu código original daqui em diante)

        # Desconto deve ser um número inteiro entre 0 e 100
        try:
            discount_percentage = int(data.get("discount", 0))
            if discount_percentage < 0:
                discount_percentage = 0
            elif discount_percentage > 100:
                discount_percentage = 100
        except (ValueError, TypeError):
            discount_percentage = 0

        # Criação da venda
        sale = Sale.objects.create(
            user=request.user,
            name=data.get("name"),
            nfe=data.get("nfe", ""),
            customer=customer,
            observations=data.get("observations", ""),
            discount=discount_percentage,
            full_price=0,
            is_quote=data.get("is_quote", False),
        )

        product_total = Decimal(0.00)

        # Processa produtos
        for item in data.get("products", []):
            product = get_object_or_404(Product, id=item["id"])
            quantity = Decimal(item["quantity"])
            price = Decimal(item["price"])

            SaleProduct.objects.create(
                sale=sale,
                product=product,
                quantity=quantity,
                price=price
            )

            subtotal = price * quantity
            product_total += subtotal

            if not sale.is_quote:
                stock = Stock.objects.filter(product=product).first()
                if stock:
                    stock.quantity -= quantity
                    stock.save()

        total_payment_before_discount = Decimal(0.00)

        # Processa pagamentos
        for payment in data.get("payments", []):
            payment_type = payment.get("payment_type")

            if payment_type != "BO":
                if payment_type == "CR":
                    installments = int(payment.get("credit_installments", 1))
                    per_installment = Decimal(payment.get("amount", 0))
                    amount = per_installment * installments
                else:
                    amount = Decimal(payment.get("amount", 0))

                SalePayment.objects.create(
                    sale=sale,
                    payment_type=payment_type,
                    amount=amount,
                    credit_installments=int(payment.get("credit_installments")) if payment_type == "CR" else None
                )

                total_payment_before_discount += amount

            else:
                boletos = payment.get("boletos", [])
                for b in boletos:
                    value = Decimal(b["value"])
                    due_date = b["due_date"]

                    Boleto.objects.create(
                        sale=sale,
                        value=value,
                        due_date=due_date,
                        status="Pendente"
                    )
                    total_payment_before_discount += value

        discount_value = (total_payment_before_discount * Decimal(discount_percentage)) / Decimal(100)
        discounted_total = total_payment_before_discount - discount_value

        sale.full_price = discounted_total
        sale.save()

        warning = None
        if product_total != discounted_total:
            warning = (
                f"Atenção: o valor total dos produtos é R$ {product_total:.2f}, "
                f"mas o valor final da venda (após desconto de {discount_percentage}%) é R$ {discounted_total:.2f}."
            )

        return JsonResponse({
            "success": True,
            "sale_id": sale.id,
            "product_total": float(product_total),
            "discount_value": float(discount_value),
            "discounted_total": float(discounted_total),
            "payment_total": float(total_payment_before_discount),
            "warning": warning
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)

@csrf_exempt
@login_required
def convert_quote_to_sale(request, sale_id):
    try:
        sale = get_object_or_404(Sale, id=sale_id)

        if not sale.is_quote:
            return JsonResponse({"success": False, "message": "Essa venda já foi concretizada."}, status=400)

        sale.is_quote = False  # Agora é uma venda
        sale.save()

        produtos_removidos = []  # Lista para armazenar produtos deletados

        # Atualizar o estoque agora que virou venda
        for sale_product in sale.sale_products.all():
            if not sale_product.product:  # O produto foi deletado
                produtos_removidos.append(sale_product.product_info)
                continue  # Pula este item e evita erro

            stock = sale_product.product.stocks.first()
            if stock:
                stock.quantity -= sale_product.quantity
                stock.save(update_fields=["quantity"])

        # Construir mensagem final
        if produtos_removidos:
            msg = "Orçamento convertido em venda, mas alguns produtos ja foram removidos: " + ", ".join(produtos_removidos)
        else:
            msg = "Orçamento convertido em venda com sucesso!"

        return JsonResponse({"success": True, "message": msg})
    
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao converter: {str(e)}"}, status=400)

@login_required
def add_manufacturer(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()

        if not name:
            return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)

        try:
            manufacturer = Manufacturer.objects.create(
                name=name
            )
            return redirect('view_manufactures')
        except Exception as e:
            return JsonResponse({"error": f"Erro ao salvar fornecedor: {str(e)}"}, status=500)
    else:
        return redirect('home')

@login_required
def add_category(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()

        if not name:
            return JsonResponse({"error": "O campo 'Nome' é obrigatório."}, status=400)

        try:
            category = Category.objects.create(
                name=name
            )
            return redirect('view_categorys')
        except Exception as e:
            return JsonResponse({"error": f"Erro ao salvar fornecedor: {str(e)}"}, status=500)
    else:
        return redirect('view_categorys')
  
@csrf_exempt
@login_required
def edit_quote(request, quote_id):
    sale = get_object_or_404(Sale, id=quote_id)

    if request.method == "POST":
        try:
            data = json.loads(request.body)

            customer_id = data.get("customer")
            customer = get_object_or_404(Customer, id=customer_id)

            sale.name = data.get("name")
            sale.nfe = data.get("nfe", "")
            sale.customer = customer
            sale.observations = data.get("observations", "")
            sale.discount = Decimal(data.get("discount", 0))
            sale.is_quote = data.get("is_quote", False)

            # Limpa dados anteriores
            sale.sale_products.all().delete()
            sale.payments.all().delete()
            sale.boletos.all().delete()

            # 🔹 Produtos
            for item in data.get("products", []):
                product = get_object_or_404(Product, id=item["id"])
                quantity = Decimal(item["quantity"])
                price = Decimal(item["price"])
                SaleProduct.objects.create(
                    sale=sale,
                    product=product,
                    quantity=quantity,
                    price=price
                )

            # 🔹 Pagamentos
            gross_payment_total = Decimal(0)
            for p in data.get("payments", []):
                payment_type = p.get("payment_type")
                amount = Decimal(p.get("amount")) if p.get("amount") else Decimal(0)
                installments = int(p.get("credit_installments")) if p.get("credit_installments") else None

                if payment_type != "BO":
                    SalePayment.objects.create(
                        sale=sale,
                        payment_type=payment_type,
                        amount=amount,
                        credit_installments=installments
                    )
                    gross_payment_total += amount
                else:
                    for boleto in p.get("boletos", []):
                        value = Decimal(boleto["value"])
                        due_date = boleto["due_date"]
                        Boleto.objects.create(
                            sale=sale,
                            value=value,
                            due_date=due_date,
                            status="Pendente"
                        )
                        gross_payment_total += value

            # 🔹 Aplica desconto ao total dos pagamentos
            discount_percentage = min(max(int(data.get("discount", 0)), 0), 100)
            discount_value = (gross_payment_total * Decimal(discount_percentage)) / Decimal(100)
            final_total = gross_payment_total - discount_value

            sale.full_price = final_total
            sale.save()

            return JsonResponse({
                "success": True,
                "total_before_discount": float(gross_payment_total),
                "discount_value": float(discount_value),
                "final_total": float(final_total)
            })

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    # Método GET - renderiza o formulário de edição
    return render(request, "pages/edit_quote.html", {"quote": sale})


@csrf_exempt
@login_required
def update_quote(request, quote_id):
    if request.method == "PUT":
        try:
            quote = get_object_or_404(Sale, id=quote_id, is_quote=True)
            data = json.loads(request.body)

            # Atualiza os campos básicos do orçamento
            quote.name = data.get("name", quote.name)
            quote.nfe = data.get("nfe", quote.nfe)
            quote.customer = get_object_or_404(Customer, id=data.get("customer"))
            quote.payment_type = data.get("payment_type", quote.payment_type)
            quote.observations = data.get("observations", quote.observations)
            quote.is_quote = data.get("is_quote", quote.is_quote)
            quote.full_price = Decimal(data.get("full_price", quote.full_price))
            quote.save()

            # Atualiza os produtos do orçamento
            quote.sale_products.all().delete()  # Remove os produtos antigos
            for product in data.get("products", []):
                product_obj = get_object_or_404(Product, id=product["id"])
                SaleProduct.objects.create(
                    sale=quote,
                    product=product_obj,
                    quantity=Decimal(product["quantity"]),
                    price=Decimal(product["price"]),
                )

            # Atualiza os boletos (se houver)
            if data.get("payment_type") == "BO":
                quote.boletos.all().delete()  # Remove os boletos antigos
                for installment in data.get("installments", []):
                    Boleto.objects.create(
                        sale=quote,
                        value=Decimal(installment["value"]),
                        due_date=installment["due_date"],
                        status="Pendente",
                    )

            return JsonResponse({"success": True, "message": "Orçamento atualizado com sucesso!"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Método não permitido"}, status=405)

@csrf_exempt
@login_required
def delete_sale(request, sale_id):
    if request.method == "DELETE":
        try:
            sale = get_object_or_404(Sale, id=sale_id)

            # Reembolsar o estoque
            for sale_product in sale.sale_products.all():
                if sale_product.product:
                    stock = Stock.objects.filter(product=sale_product.product).first()
                    if stock:
                        stock.quantity += sale_product.quantity
                        stock.save(update_fields=["quantity"])

            sale.delete()
            return JsonResponse({"success": True, "message": "Venda deletada com sucesso."}, status=200)

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Método não permitido"}, status=405)

@csrf_exempt
@login_required
def add_purchase(request):
    if request.method == "POST":
        try:
            # Parse os dados JSON enviados
            data = json.loads(request.body)

            # Captura e valida o fornecedor (supplier)
            supplier_id = data.get("supplier")
            supplier = get_object_or_404(Supplier, id=supplier_id)

            # Captura o usuário (ajuste conforme sua lógica de autenticação)
            user = request.user

            # Inicializa o valor total da compra
            full_price = Decimal(0.00)

            # Cria a compra dentro de uma transação
            with transaction.atomic():
                purchase = Purchase.objects.create(
                    user=user,
                    name=data.get("name"),
                    supplier=supplier,
                    observations=data.get("observations", ""),
                    full_price=full_price  # Atualizaremos depois de calcular o preço total
                )

                # Processa os produtos
                products = data.get("products", [])
                for item in products:
                    product = get_object_or_404(Product, id=item.get("id"))
                    quantity = Decimal(item.get("quantity", 1))

                    # Verifica o preço do produto no estoque
                    stock, created = Stock.objects.get_or_create(product=product)

                    # Atualiza o estoque (incrementa a quantidade)
                    stock.quantity += quantity
                    stock.save()

                    # Calcula o total do item
                    # Captura o preço informado pelo usuário ou usa o do estoque como fallback
                    custom_price = Decimal(item.get("price", stock.price))  

                    # Calcula o total do item
                    item_total = custom_price * quantity
                    full_price += item_total
                    
                    # Cria o PurchaseProduct
                    PurchaseProduct.objects.create(
                        purchase=purchase,
                        product=product,
                        quantity=quantity,
                        price=custom_price
                    )

                # Atualiza o preço total da compra
                purchase.full_price = full_price
                purchase.save()

            # Retorna resposta de sucesso
            return JsonResponse({"success": True, "message": "Compra registrada com sucesso!"}, status=201)

        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=400)

    return JsonResponse({"success": False, "message": "Método não permitido."}, status=405)

@csrf_exempt
@login_required
def delete_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)

    for purchase_product in purchase.purchase_products.all():
        stock, created = Stock.objects.get_or_create(product=purchase_product.product)
        
        # Calcula o total já vendido desse produto
        total_sold = sum(sp.quantity for sp in purchase_product.product.sale_products.all())

        # Define a nova quantidade do estoque
        new_quantity = stock.quantity - purchase_product.quantity

        # Garante que o estoque nunca fique menor que o total vendido
        stock.quantity = max(new_quantity, -total_sold)  
        stock.save()

    purchase.delete()
    messages.success(request, "Compra deletada e estoque ajustado corretamente.")
    return redirect("view_purchases")

@csrf_exempt
@login_required
def delete_purchase(request, purchase_id):
    if request.method == "DELETE":
        try:
            purchase = get_object_or_404(Purchase, id=purchase_id)

            # Reembolsar o estoque
            for purchase_product in purchase.purchase_products.all():
                if purchase_product.product:
                    stock = Stock.objects.filter(product=purchase_product.product).first()
                    if stock:
                        stock.quantity += purchase_product.quantity
                        stock.save(update_fields=["quantity"])

            purchase.delete()
            return JsonResponse({"success": True, "message": "Compra deletada com sucesso."}, status=200)

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Método não permitido"}, status=405)

@csrf_exempt
@login_required
def update_stock(request, stock_id):
    if request.method == "POST":
        try:
            stock = get_object_or_404(Stock, id=stock_id)
            data = json.loads(request.body)

            # Atualiza a quantidade
            new_quantity = data.get("quantity")
            if new_quantity is not None:
                stock.quantity = new_quantity

            # Atualiza o preço 
            new_price = data.get("price")
            if new_price is not None:
                stock.price = new_price

            # Atualiza a quantidade mínima
            new_min_quantity = data.get("min_quantity")
            if new_min_quantity is not None:
                stock.min_quantity = new_min_quantity

            # Atualiza a quantidade máxima
            new_max_quantity = data.get("max_quantity")
            if new_max_quantity is not None:
                stock.max_quantity = new_max_quantity

            stock.save()

            return JsonResponse({"success": True, "message": "Estoque atualizado com sucesso."})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "message": "Método não permitido."}, status=405)

@csrf_exempt
@login_required
def delete_product(request, product_id):
    """
    View para deletar um produto pelo ID.
    Aceita apenas requisições DELETE.
    """
    if request.method == "POST":
        try:
            product = get_object_or_404(Product, id=product_id)
            product.delete()
            return JsonResponse({"message": "Produto deletado com sucesso."}, status=200)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Método não permitido."}, status=405)

def pay_boleto(request, boleto_id):
    if request.method == "POST":
        boleto = get_object_or_404(Boleto, id=boleto_id)
        boleto.status = "Pago"
        boleto.save()
        return JsonResponse({"message": "Boleto marcado como pago com sucesso.", "boleto_id": boleto.id})
    return JsonResponse({"error": "Método não permitido."}, status=405)

# ======================== RELATORIOS =======================================

# ============ PRODUTOS
def gerar_relatorio_produtos(request):
    # Recebe o filtro de categoria e formato
    category_id = request.GET.get('categoria')  # ID da categoria recebida via GET
    formato = request.GET.get('formato', 'excel')  # Filtro de formato, padrão é 'excel'

    # Filtra a categoria, se fornecido
    categoria = None
    if category_id:
        try:
            categoria = Category.objects.get(id=category_id)  # Obtém a categoria pelo ID
        except Category.DoesNotExist:
            categoria = None

    # Recupera os dados da API (simulando o retorno da API para os produtos)
    stock_data = fetch_stock_data()  # Supondo que você tenha uma função que busca dados da API

    # Filtra os dados com base na categoria, se fornecido
    if categoria:
        stock_data = [stock for stock in stock_data if stock['product']['category'] == categoria.name]

    # Chama a função de acordo com o formato escolhido (excel ou pdf)
    if formato == 'excel':
        return gerar_relatorio_produtos_excel(stock_data, categoria)
    elif formato == 'pdf':
        return gerar_relatorio_produtos_pdf(stock_data, categoria)
    else:
        return JsonResponse({"success": False, "error": "Formato não suportado"}, status=400)

def gerar_relatorio_produtos_excel(request, categoria):
    # Busca os dados da API
    stock_data = fetch_stock_data()

    # Filtra por categoria, se fornecida
    if categoria:
        stock_data = [stock for stock in stock_data if stock['product']['category'] == categoria.name]

    # Criação do arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Produtos"

    # Título do relatório
    titulo = f"Relatório de Produtos - Categoria: {categoria.name if categoria else 'Todas'}"
    ws.merge_cells('A1:J1')
    ws['A1'] = titulo
    ws['A1'].alignment = Alignment(horizontal='center')

    # Cabeçalhos da tabela
    headers = [
        "Nome do Produto", "Gramatura", "Formato", "Quantidade de Pacotes",
        "Quantidade por Pacote", "Total de Folhas", "Preço Unitário",
        "Preço do KG", "Peso do Pacote", "Peso Total"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 2

    # Preenche os dados da tabela
    for row_num, stock in enumerate(stock_data, start=3):
        product = stock['product']

        # Convertendo os valores numéricos para o tipo adequado
        quantity = int(float(stock['quantity']))
        price = float(stock['price'])
        total_folhas = int(float(stock['total_folhas']))
        preco_por_kg = float(stock['preco_por_kg'])
        peso_por_pacote = float(stock['peso_por_pacote'])
        peso_total = float(stock['peso_total'])
        preco_total = float(stock['preco_total'])

        ws.cell(row=row_num, column=1, value=product['name'])
        ws.cell(row=row_num, column=2, value=f"{product['weight']} {product['size']}")
        ws.cell(row=row_num, column=3, value=product['size'])
        ws.cell(row=row_num, column=4, value=quantity)
        ws.cell(row=row_num, column=5, value=int(product['quantity_per_package']))
        ws.cell(row=row_num, column=6, value=total_folhas)
        ws.cell(row=row_num, column=7, value=f"R$ {price:.2f}")
        ws.cell(row=row_num, column=8, value=f"R$ {preco_por_kg:.2f}")
        ws.cell(row=row_num, column=9, value=f"{peso_por_pacote:.2f} kg")
        ws.cell(row=row_num, column=10, value=f"{peso_total:.2f} kg")

        # Alinhamento à esquerda para todas as células
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left')

    # Retorno do arquivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"relatorio_produtos_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def gerar_relatorio_produtos_pdf(stock_data, categoria):
    # Criação do arquivo PDF
    buffer = BytesIO()
    c = SimpleDocTemplate(buffer, pagesize=A4)

    # Título do relatório
    titulo = f"Relatório de Produtos - Categoria: {categoria.name if categoria else 'Todas'}"

    # Usando o Paragraph para título
    styles = getSampleStyleSheet()
    title_paragraph = Paragraph(titulo, styles['Title'])
    
    # Adicionando título ao PDF
    story = [title_paragraph]

    # Cabeçalhos da tabela com quebras de linha e abreviações
    headers = [
        "Nome\nProduto", "Gram.\n(kg/g)", "Formato", "Pacotes",
        "Qtd./\nPacote", "Total\nFolhas", "Preço\nUnit.", "Preço\nKG",
        "Peso\nPacote", "Peso\nTotal"
    ]
    
    # Dados dos produtos
    data = [headers]
    for stock in stock_data:
        produto = stock['product']

        # Quebra o nome em linhas de até 25 caracteres e limita a 2 linhas
        nome_formatado = "\n".join(wrap(produto['name'], width=25)[:2])
        if len(wrap(produto['name'], width=25)) > 2:
            nome_formatado += "..."  # Adiciona reticências se for truncado
        
        preco_unitario = float(stock['preco_unitario']) if stock.get('preco_unitario') else 0
        preco_por_kg = float(stock['preco_por_kg']) if stock.get('preco_por_kg') else 0
        peso_por_pacote = float(stock['peso_por_pacote']) if stock.get('peso_por_pacote') else 0
        peso_total = float(stock['peso_total']) if stock.get('peso_total') else 0

        # Preenche as células com os valores calculados
        row = [
            nome_formatado,
            f"{produto['weight']} {produto['unit_type']}",
            produto['size'],
            stock['quantity'],
            produto['quantity_per_package'],
            stock['total_folhas'],
            f"R$ {float(preco_unitario):.2f}",
            f"R$ {float(preco_por_kg):.2f}",
            f"{float(peso_por_pacote):.2f} kg",
            f"{float(peso_total):.2f} kg",
        ]
        data.append(row)

    # Criando a tabela com os dados
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ])
    table.setStyle(style)

    # Adiciona a tabela ao PDF
    story.append(table)

    # Finaliza a criação do PDF
    c.build(story)

    # Retorna o PDF como resposta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_produtos.pdf"'
    return response

# ============ VENDAS
def gerar_relatorio_vendas(request):
    # Recebe as datas de início e fim do filtro
    data_inicial = request.GET.get('data_inicial')
    data_final = request.GET.get('data_final')
    formato = request.GET.get('formato', 'excel')  # Default para 'excel'

    # Filtro de vendas
    vendas = Sale.objects.all()

    # Filtra pelo período, se as datas forem fornecidas
    if data_inicial:
        data_inicial_obj = datetime.strptime(data_inicial, "%Y-%m-%d").date()
        vendas = vendas.filter(created__date__gte=data_inicial_obj)

    if data_final:
        data_final_obj = datetime.strptime(data_final, "%Y-%m-%d").date()
        vendas = vendas.filter(created__date__lte=data_final_obj)

    if data_inicial and not data_final:
        vendas = vendas.filter(created__date=data_inicial_obj)

    if formato == 'excel':
        return gerar_excel_vendas(vendas, data_inicial, data_final)
    elif formato == 'pdf':
        return gerar_pdf_vendas(vendas, data_inicial, data_final)
    else:
        return JsonResponse({"success": False, "error": "Formato não suportado"}, status=400)

def gerar_excel_compras(compras, data_inicial, data_final):
    # Função para gerar o relatório em Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Compras"
    
    # Título do relatório
    if data_inicial and data_final:
        filtro_data = f"De {data_inicial} a {data_final}"
    elif data_inicial:
        filtro_data = f"Somente o dia {data_inicial}"  # Ajuste para o caso de somente a data inicial ser fornecida
    else:
        filtro_data = "Sem filtro de data"

    titulo = f"Relatório de Compras - Filtro: {filtro_data}"
    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].alignment = Alignment(horizontal='center')

    # Cabeçalhos da tabela
    headers = [
        "ID da Compra", "Fornecedor", "Usuario", "Data da Compra", "Valor Total", "Total de Itens", "Observações"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 2

    # Dados das compras
    for row_num, compra in enumerate(compras, start=3):
        total_itens = sum(purchase_product.quantity for purchase_product in compra.purchase_products.all())  # Soma das quantidades de produtos vendidos
        ws.cell(row=row_num, column=1, value=compra.id)
        ws.cell(row=row_num, column=2, value=compra.supplier.name)
        ws.cell(row=row_num, column=3, value=compra.user.username)
        ws.cell(row=row_num, column=4, value=compra.created.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=6, value=compra.full_price)
        ws.cell(row=row_num, column=7, value=total_itens)
        ws.cell(row=row_num, column=8, value=compra.observations or "")

        # Alinhamento à esquerda
        for col_num in range(1, 9):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left')

    # Retorno do arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"relatorio_compras_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def gerar_pdf_vendas(vendas, data_inicial, data_final):
    # Criação do arquivo PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Título do relatório
    if data_inicial and data_final:
        filtro_data = f"De {data_inicial} a {data_final}"
    elif data_inicial:
        filtro_data = f"Somente o dia {data_inicial}"
    else:
        filtro_data = "Sem filtro de data"

    titulo = f"Relatório de Vendas - Filtro: {filtro_data}"
    styles = getSampleStyleSheet()
    title_paragraph = Paragraph(titulo, styles['Title'])
    elements.append(title_paragraph)

    # Cabeçalhos da tabela
    headers = [
        "ID", "Cliente", "Vendedor", "Data da Venda",
        "Pagamento", "Valor Total", "Total de Itens"
    ]

    # Dados das vendas
    data = [headers]
    for venda in vendas:
        total_itens = sum(sale_product.quantity for sale_product in venda.sale_products.all())
        row = [
            venda.id,
            venda.customer.name,
            venda.user.username,
            venda.created.strftime('%d/%m/%Y'),
            venda.get_payment_type_display(),
            f"R$ {venda.full_price:.2f}",
            total_itens,
        ]
        data.append(row)

    # Criação da tabela
    table = Table(data, colWidths=[30, 100, 80, 80, 80, 80, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ])
    table.setStyle(style)

    # Adiciona a tabela ao PDF
    elements.append(table)

    # Geração do PDF
    doc.build(elements)
    buffer.seek(0)

    # Retorna o PDF como resposta HTTP
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ============ COMPRAS
def gerar_relatorio_compras(request):
    # Recebe as datas de início e fim do filtro
    data_inicial = request.GET.get('data_inicial')
    data_final = request.GET.get('data_final')
    formato = request.GET.get('formato', 'excel')  # Default para 'excel'

    # Filtro de compras
    compras = Purchase.objects.all()

    # Filtra pelo período, se as datas forem fornecidas
    if data_inicial:
        data_inicial_obj = datetime.strptime(data_inicial, "%Y-%m-%d").date()
        compras = compras.filter(created__date__gte=data_inicial_obj)

    if data_final:
        data_final_obj = datetime.strptime(data_final, "%Y-%m-%d").date()
        compras = compras.filter(created__date__lte=data_final_obj)

    if data_inicial and not data_final:
        compras = compras.filter(created__date=data_inicial_obj)

    if formato == 'excel':
        return gerar_excel_compras(compras, data_inicial, data_final)
    elif formato == 'pdf':
        return gerar_pdf_compras(compras, data_inicial, data_final)
    else:
        return JsonResponse({"success": False, "error": "Formato não suportado"}, status=400)

def gerar_excel_compras(compras, data_inicial, data_final):
    # Função para gerar o relatório em Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Compras"
    
    # Título do relatório
    if data_inicial and data_final:
        filtro_data = f"De {data_inicial} a {data_final}"
    elif data_inicial:
        filtro_data = f"Somente o dia {data_inicial}"  # Ajuste para o caso de somente a data inicial ser fornecida
    else:
        filtro_data = "Sem filtro de data"

    titulo = f"Relatório de Compras - Filtro: {filtro_data}"
    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].alignment = Alignment(horizontal='center')

    # Cabeçalhos da tabela
    headers = [
        "ID da Compra", "Fornecedor", "Usuario", "Data da Compra", "Valor Total", "Total de Itens", "Observações"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 2

    # Dados das compras
    for row_num, compra in enumerate(compras, start=3):
        total_itens = sum(purchase_product.quantity for purchase_product in compra.purchase_products.all())  # Soma das quantidades de produtos vendidos
        ws.cell(row=row_num, column=1, value=compra.id)
        ws.cell(row=row_num, column=2, value=compra.supplier.name)
        ws.cell(row=row_num, column=3, value=compra.user.username)
        ws.cell(row=row_num, column=4, value=compra.created.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=6, value=compra.full_price)
        ws.cell(row=row_num, column=7, value=total_itens)
        ws.cell(row=row_num, column=8, value=compra.observations or "")

        # Alinhamento à esquerda
        for col_num in range(1, 9):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left')

    # Retorno do arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"relatorio_compras_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def gerar_pdf_compras(compras, data_inicial, data_final):
    # Criação do arquivo PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Título do relatório
    if data_inicial and data_final:
        filtro_data = f"De {data_inicial} a {data_final}"
    elif data_inicial:
        filtro_data = f"Somente o dia {data_inicial}"
    else:
        filtro_data = "Sem filtro de data"

    titulo = f"Relatório de Compras - Filtro: {filtro_data}"
    styles = getSampleStyleSheet()
    title_paragraph = Paragraph(titulo, styles['Title'])
    elements.append(title_paragraph)

    # Cabeçalhos da tabela
    headers = [
        "ID", "Fornecedor", "Usuário", "Data da Compra",
        "Valor Total", "Total de Itens"
    ]

    # Dados das compras
    data = [headers]
    for compra in compras:
        total_itens = sum(purchase_product.quantity for purchase_product in compra.purchase_products.all())
        row = [
            compra.id,
            compra.supplier.name,
            compra.user.username,
            compra.created.strftime('%d/%m/%Y'),
            f"R$ {compra.full_price:.2f}",
            total_itens,
        ]
        data.append(row)

    # Criação da tabela
    table = Table(data, colWidths=[20, 100, 80, 80, 80, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ])
    table.setStyle(style)

    # Adiciona a tabela ao PDF
    elements.append(table)

    # Geração do PDF
    doc.build(elements)
    buffer.seek(0)

    # Retorna o PDF como resposta HTTP
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"relatorio_compras_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ======================= API =================================================

def get_all_categories(request):
    categories = Category.objects.all()
    categories_list = [{
        "id": category.id,
        'name': category.name
    } for category in categories]

    return JsonResponse({'categories': categories_list})

def get_all_manufacturers(request):
    manufacturers = Manufacturer.objects.all()
    manufacturers_list = [{
        "id": manufacturer.id,
        'name': manufacturer.name
    } for manufacturer in manufacturers]

    return JsonResponse({'manufacturers': manufacturers_list})

def get_all_categorys(request):
    category = Category.objects.all()
    category_list = [{
        "id": category.id,
        'name': category.name
    } for category in category]

    return JsonResponse({'category': category_list})

def get_all_customers(request):
    query = request.GET.get("search", "").strip()

    customers = Customer.objects.all()

    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(cpf_or_cnpj__icontains=query)
        )

    # Busca clientes com boletos vencidos
    today = now().date()
    boletos_vencidos = Boleto.objects.filter(
        status="Pendente",
        due_date__lt=today
    )
    clientes_com_vencidos = set(
        boletos_vencidos.values_list("sale__customer_id", flat=True)
    )

    customer_list = [{
        'id': customer.id,
        'name': customer.name,
        'cpf_or_cnpj': customer.cpf_or_cnpj,
        'is_cnpj': customer.is_cnpj,
        'street': customer.street,
        'address_number': customer.address_number,
        'district': customer.district,
        'city': customer.city,
        'state': customer.state,
        'cep': customer.cep,
        'contact_name': customer.contact_name,
        'phone': customer.phone,
        'email': customer.email,
        'observations': customer.observations,
        'has_overdue_boleto': customer.id in clientes_com_vencidos,
    } for customer in customers]

    return JsonResponse({'customers': customer_list})

def get_all_suppliers(request):
    query = request.GET.get("search", "").strip()

    suppliers = Supplier.objects.all()

    if query:
        suppliers = suppliers.filter(
            Q(name__icontains=query) |
            Q(cnpj__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query)
        )

    supplier_list = [{
        'id': supplier.id,
        'name': supplier.name,
        'contact_name': supplier.contact_name,
        'phone': supplier.phone,
        'email': supplier.email,
        'street': supplier.street,
        'address_number': supplier.address_number,
        'district': supplier.district,
        'city': supplier.city,
        'state': supplier.state,
        'cep': supplier.cep,
        'cnpj': supplier.cnpj,
        'observations': supplier.observations,
        'created': supplier.created.strftime('%Y-%m-%d %H:%M:%S'),
    } for supplier in suppliers]

    return JsonResponse({'suppliers': supplier_list})

def get_all_products(request):
    query = request.GET.get("search", "").strip()

    products = Product.objects.all()

    if query:
        products = products.filter(Q(name__icontains=query))

    product_list = [{
        'id': product.id,
        'name': product.name,
        'manufacturer': {
            'id': product.manufacturer.id,
            'name': product.manufacturer.name
        } if product.manufacturer else None,
        'category': {
            'id': product.category.id,
            'name': product.category.name
        } if product.category else None,
        'weight': product.weight,
        'unit_type': product.get_unit_type_display(),
        'size': product.size,
        'quantity_per_package': product.quantity_per_package,
    } for product in products]

    return JsonResponse({'products': product_list})


def api_products(request):
    query = request.GET.get('search', '')

    if query:
        products = Product.objects.filter(
            Q(name__icontains=query) |
            Q(category__name__icontains=query) |
            Q(manufacturer__name__icontains=query)
        )
    else:
        products = Product.objects.all()

    # Anotando com menor preço do estoque
    products_with_price = products.annotate(min_price=Min('stocks__price'))

    data = {
        "products": [
            {
                "id": product.id,
                "name": product.name,
                "category": product.category.name if product.category else "",
                "manufacturer": product.manufacturer.name if product.manufacturer else "",
                "weight": product.weight,
                "unit_type": product.unit_type,
                "size": product.size,
                "quantity_per_package": product.quantity_per_package,
                "price": product.min_price if product.min_price is not None else 0
            }
            for product in products_with_price
        ]
    }

    return JsonResponse(data)

def get_unit_types(request):
    unit_types = [{"value": key, "label": value} for key, value in dict(Product.UNIT_CHOICES).items()]
    return JsonResponse({"unit_types": unit_types})

def get_all_stocks(request):
    query = request.GET.get("search", "").strip()

    stocks = Stock.objects.select_related("product").all()

    if query:
        stocks = stocks.filter(Q(product__name__icontains=query))  # Filtra pelo nome do produto

    stock_list = []

    for stock in stocks:
        price = float(stock.price) if stock.price else 0.0
        quantity = float(stock.quantity) if stock.quantity else 0.0  # Converte para float
        peso_por_pacote = float(getattr(stock, "peso_por_pacote", 0.0))  # Converte para float
        preco_por_kg = float(stock.preco_por_kg) if hasattr(stock, "preco_por_kg") else (price / peso_por_pacote if peso_por_pacote else 0.0)
        peso_total = peso_por_pacote * quantity  # Ambos são float agora
        preco_total = price * quantity  # Ambos são float agora

        stock_list.append({
            "id": stock.id,
            "product_id": stock.product.id if stock.product else None,
            "product_name": stock.product.name if stock.product else "Produto Removido",
            "quantity": quantity,
            "price": price,
            "preco_por_kg": preco_por_kg,
            "peso_por_pacote": peso_por_pacote,
            "peso_total": peso_total,
            "preco_total": preco_total,
            "min_quantity": float(stock.min_quantity) if stock.min_quantity else 0.0,
            "max_quantity": float(stock.max_quantity) if stock.max_quantity else 0.0,
        })

    return JsonResponse({"stocks": stock_list})

def api_stock(request):
    query = request.GET.get('search', '')

    # Filtra os estoques com base no nome do produto
    if query:
        stocks = Stock.objects.filter(product__name__icontains=query)
    else:
        stocks = Stock.objects.all()

    # Retorna os dados do estoque em JSON
    data = {
        "stocks": [
            {
                "id": stock.id,
                "product_name": stock.product.name,
                "quantity": float(stock.quantity),  # Garante que seja um número válido no JSON
                "price": float(stock.price) if stock.price else 0,
                "preco_por_kg": float(stock.preco_por_kg),
                "preco_total": float(stock.preco_total),
                "peso_por_pacote": float(stock.peso_por_pacote),
                "peso_total": float(stock.peso_total),
                "min_quantity": float(stock.min_quantity),  # Adiciona min_quantity
                "max_quantity": float(stock.max_quantity),  # Adiciona max_quantity
            }
            for stock in stocks
        ]
    }
    return JsonResponse(data)

def fetch_stock_data():
    response = requests.get("http://127.0.0.1:8000/api/stocks/")
    if response.status_code == 200:
        return response.json().get('stock', [])
    else:
        return []

def get_all_sales(request):
    query = request.GET.get("search", "").strip()
    sales = Sale.objects.all()

    if query:
        sales = sales.filter(
            Q(name__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(nfe__icontains=query)
        )

    sale_list = []

    for sale in sales:
        sale_products = [
            {
                "product_id": sp.product.id if sp.product else None,
                "product_name": sp.product.name if sp.product else sp.product_info,
                "product_category": sp.product.category.name if sp.product and sp.product.category else "",
                "product_manufaturer": sp.product.manufacturer.name if sp.product and sp.product.manufacturer else "",
                "quantity": float(sp.quantity),
                "price": float(sp.price),
                "total_price": float(sp.quantity * sp.price),
            }
            for sp in sale.sale_products.all()
        ]

        payments = [
            {
                "payment_type": p.payment_type,
                "payment_type_display": p.get_payment_type_display(),
                "amount": float(p.amount),
                "credit_installments": p.credit_installments,
            }
            for p in sale.payments.all()
        ]

        boletos = [
            {
                "value": float(b.value),
                "due_date": b.due_date.strftime("%Y-%m-%d"),
                "status": b.status,
            }
            for b in sale.boletos.all()
        ]

        sale_list.append({
            "sale_id": sale.id,
            "sale_name": sale.name,
            "sale_created": sale.created.strftime("%Y-%m-%d %H:%M:%S"),
            "customer_name": sale.customer.name if sale.customer else sale.customer_name,
            "full_price": float(sale.full_price),
            "discount": float(sale.discount or 0),
            "observations": sale.observations,
            "is_quote": sale.is_quote,
            "nfe": sale.nfe,
            "user_id": sale.user.id,
            "user_name": sale.user.username,
            "products": sale_products,
            "payments": payments,
            "boletos": boletos,
        })

    return JsonResponse({"sales": sale_list})

def get_all_confirmed_sales(request):
    query = request.GET.get("search", "").strip()
    
    sales = Sale.objects.filter(is_quote=False).order_by('-created')

    if query:
        sales = sales.filter(
            Q(name__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(nfe__icontains=query)
        )

    sale_list = []

    for sale in sales:
        sale_products = [
            {
                "product_id": sp.product.id if sp.product else None,
                "product_name": sp.product.name if sp.product else sp.product_info,
                "product_category": sp.product.category.name if sp.product and sp.product.category else "",
                "product_manufaturer": sp.product.manufacturer.name if sp.product and sp.product.manufacturer else "",
                "quantity": float(sp.quantity),
                "price": float(sp.price),
                "total_price": float(sp.quantity * sp.price),
            }
            for sp in sale.sale_products.all()
        ]

        payments = [
            {
                "payment_type": p.payment_type,
                "payment_type_display": p.get_payment_type_display(),
                "amount": float(p.amount),
                "credit_installments": p.credit_installments,
            }
            for p in sale.payments.all()
        ]

        boletos = [
            {
                "value": float(b.value),
                "due_date": b.due_date.strftime("%Y-%m-%d"),
                "status": b.status,
            }
            for b in sale.boletos.all()
        ]

        sale_list.append({
            "sale_id": sale.id,
            "sale_name": sale.name,
            "sale_created": sale.created.strftime("%Y-%m-%d %H:%M:%S"),
            "customer_name": sale.customer.name if sale.customer else sale.customer_name,
            "full_price": float(sale.full_price),
            "discount": float(sale.discount or 0),
            "observations": sale.observations,
            "is_quote": sale.is_quote,
            "nfe": sale.nfe,
            "user_id": sale.user.id,
            "user_name": sale.user.username,
            "products": sale_products,
            "payments": payments,
            "boletos": boletos,
        })

    return JsonResponse({"sales": sale_list})

def get_all_purchases(request):
    query = request.GET.get("search", "").strip()

    purchases = Purchase.objects.all().order_by('-created')

    if query:
        purchases = purchases.filter(
            Q(name__icontains=query) |  # Busca pelo nome da compra
            Q(supplier__name__icontains=query)  # Busca pelo nome do fornecedor
        )

    purchase_list = []

    for purchase in purchases:
        purchase_products = [
            {
                "product_id": purchase_product.product.id if purchase_product.product else None,
                "product_name": purchase_product.product.name if purchase_product.product else purchase_product.product_info,
                "product_category": purchase_product.product.category.name if purchase_product.product and purchase_product.product.category else "",
                "product_manufaturer": purchase_product.product.manufacturer.name if purchase_product.product and purchase_product.product.manufacturer else "",
                "quantity": float(purchase_product.quantity),
                "price": float(purchase_product.price),
                "total_price": float(purchase_product.quantity * purchase_product.price),
            }
            for purchase_product in purchase.purchase_products.all()
        ]

        purchase_list.append({
            "purchase_id": purchase.id,
            "purchase_name": purchase.name,
            "purchase_created": purchase.created,
            "supplier_name": purchase.supplier.name if purchase.supplier else purchase.supplier_name,
            "full_price": float(purchase.full_price),
            "products": purchase_products,
        })

    return JsonResponse({"purchases": purchase_list})



def get_summary(request):
    tz = get_current_timezone()
    now_local = localtime()
    today = now_local.date()

    # Constrói os intervalos em horário local
    start_of_day_local = datetime.combine(today, datetime.min.time()).replace(tzinfo=tz)
    end_of_day_local = datetime.combine(today, datetime.max.time()).replace(tzinfo=tz)

    start_of_week_local = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time()).replace(tzinfo=tz)
    start_of_month_local = datetime.combine(today.replace(day=1), datetime.min.time()).replace(tzinfo=tz)

    # Converte para UTC (já que o banco armazena timestamps em UTC)
    start_of_day_utc = start_of_day_local.astimezone(timezone.utc)
    end_of_day_utc = end_of_day_local.astimezone(timezone.utc)
    start_of_week_utc = start_of_week_local.astimezone(timezone.utc)
    start_of_month_utc = start_of_month_local.astimezone(timezone.utc)

    # Filtra as vendas concretizadas (ignora orçamentos e com valor zero)
    daily_sales = Sale.objects.filter(created__range=(start_of_day_utc, end_of_day_utc), is_quote=False, full_price__gt=0)
    weekly_sales = Sale.objects.filter(created__gte=start_of_week_utc, is_quote=False, full_price__gt=0)
    monthly_sales = Sale.objects.filter(created__gte=start_of_month_utc, is_quote=False, full_price__gt=0)

    daily_purchases = Purchase.objects.filter(created__range=(start_of_day_utc, end_of_day_utc))
    weekly_purchases = Purchase.objects.filter(created__gte=start_of_week_utc)
    monthly_purchases = Purchase.objects.filter(created__gte=start_of_month_utc)

    def calculate_summary(sales, purchases):
        sales_products = SaleProduct.objects.filter(sale__in=sales)
        purchases_products = PurchaseProduct.objects.filter(purchase__in=purchases)

        sales_total = sales.aggregate(total=Sum('full_price'))['total'] or 0
        purchases_total = purchases.aggregate(total=Sum('full_price'))['total'] or 0
        products_sold = sales_products.aggregate(total=Sum('quantity'))['total'] or 0
        products_purchased = purchases_products.aggregate(total=Sum('quantity'))['total'] or 0

        profit = sales_total - purchases_total

        return {
            'sales_total': float(sales_total),
            'purchases_total': float(purchases_total),
            'products_sold': float(products_sold),
            'products_purchased': float(products_purchased),
            'profit': float(profit),
        }

    return JsonResponse({
        'daily': calculate_summary(daily_sales, daily_purchases),
        'weekly': calculate_summary(weekly_sales, weekly_purchases),
        'monthly': calculate_summary(monthly_sales, monthly_purchases),
    })
def get_latest_sales(request):
    query = request.GET.get("search", "").strip()
    
    sales = Sale.objects.filter(is_quote=False).order_by('-created')[:10]

    if query:
        sales = sales.filter(
            Q(name__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(nfe__icontains=query)
        )

    sale_list = []

    for sale in sales:
        sale_products = [
            {
                "product_id": sp.product.id if sp.product else None,
                "product_name": sp.product.name if sp.product else sp.product_info,
                "product_category": sp.product.category.name if sp.product and sp.product.category else "",
                "product_manufaturer": sp.product.manufacturer.name if sp.product and sp.product.manufacturer else "",
                "quantity": float(sp.quantity),
                "price": float(sp.price),
                "total_price": float(sp.quantity * sp.price),
            }
            for sp in sale.sale_products.all()
        ]

        payments = [
            {
                "payment_type": p.payment_type,
                "payment_type_display": p.get_payment_type_display(),
                "amount": float(p.amount),
                "credit_installments": p.credit_installments,
            }
            for p in sale.payments.all()
        ]

        boletos = [
            {
                "value": float(b.value),
                "due_date": b.due_date.strftime("%Y-%m-%d"),
                "status": b.status,
            }
            for b in sale.boletos.all()
        ]

        sale_list.append({
            "sale_id": sale.id,
            "sale_name": sale.name,
            "sale_created": sale.created.strftime("%Y-%m-%d %H:%M:%S"),
            "customer_name": sale.customer.name if sale.customer else sale.customer_name,
            "full_price": float(sale.full_price),
            "discount": float(sale.discount or 0),
            "observations": sale.observations,
            "is_quote": sale.is_quote,
            "nfe": sale.nfe,
            "user_id": sale.user.id,
            "user_name": sale.user.username,
            "products": sale_products,
            "payments": payments,
            "boletos": boletos,
        })

    return JsonResponse({"sales": sale_list})

def get_latest_purchases(request):
    purchases = Purchase.objects.order_by('-created')[:10]
    data = {
        'purchases': [
            {
                'purchase_id': purchase.id,
                'purchase_name': purchase.name,
                'supplier_name': purchase.supplier.name if purchase.supplier else purchase.supplier_name,  # CORRIGIDO
                'purchase_created': purchase.created,
                'full_price': float(purchase.full_price),
            }
            for purchase in purchases
        ]
    }
    return JsonResponse(data)

def get_boletos_pendentes(request):
    # Filtrar apenas os boletos com status "Pendente"
    boletos_pendentes = Boleto.objects.filter(status="Pendente")

    # Pré-carregar os dados relacionados para evitar consultas extras
    boletos = boletos_pendentes.select_related("sale__customer", "sale__user")

    today = now().date()

    # Construir a lista de boletos
    boletos_list = [
        {
            "boleto_id": boleto.id,
            "valor": f"R$ {boleto.value:.2f}",
            "data_vencimento": boleto.due_date.strftime('%d/%m/%Y'),
            "vencido": boleto.due_date < today,  # alerta de vencimento
            "venda": {
                "venda_id": boleto.sale.id,
                "nome": boleto.sale.name,
                "nfe": boleto.sale.nfe,
                "data_venda": boleto.sale.created.strftime('%d/%m/%Y'),
                "cliente": {
                    "id": boleto.sale.customer.id,
                    "nome": boleto.sale.customer.name,
                },
                "usuario": {
                    "id": boleto.sale.user.id,
                    "nome": boleto.sale.user.username,
                },
                "valor_total": f"R$ {boleto.sale.full_price:.2f}",
                "observacoes": boleto.sale.observations or "",
            },
        }
        for boleto in boletos
    ]

    return JsonResponse({"boletos_pendentes": boletos_list})

def get_negative_stocks(request):
    """
    Retorna todos os estoques com quantidade negativa.
    """
    negative_stocks = Stock.objects.filter(quantity__lt=0).select_related('product')
    stock_list = []

    for stock in negative_stocks:
        product = stock.product
        stock_list.append({
            'product_id': product.id,
            'product_name': product.name,
            'category': product.category.name if product.category else "Sem categoria",
            'manufacturer': product.manufacturer.name if product.manufacturer else "Sem Fabricante",
            'quantity': float(stock.quantity),  # Convertendo para evitar problemas com JSON
            'price': float(stock.price),
            'pending': float(stock.pending),
        })

    return JsonResponse({'negative_stocks': stock_list})

def get_all_quotes(request):
    quotes = Sale.objects.filter(is_quote=True)  # Apenas orçamentos
    quote_list = [{
        "id": quote.id,
        "name": quote.name,
        "customer_name": quote.customer_name,
        "full_price": float(quote.full_price),
        "created": quote.created,
    } for quote in quotes]

    return JsonResponse({"quotes": quote_list})

@login_required
def get_quote_data(request, quote_id):
    try:
        sale = Sale.objects.get(id=quote_id)
        products = sale.sale_products.all()
        payments = sale.payments.all()
        boletos = sale.boletos.all()

        return JsonResponse({
            "success": True,
            "quote": {
                "id": sale.id,
                "name": sale.name,
                "nfe": sale.nfe,
                "customer": sale.customer_id,
                "observations": sale.observations,
                "discount": int(sale.discount or 0),
                "is_quote": sale.is_quote,
                "products": [
                    {
                        "id": sp.product.id if sp.product else None,
                        "name": sp.product.name if sp.product else sp.product_info,
                        "quantity": float(sp.quantity),
                        "price": float(sp.price),
                    }
                    for sp in products
                ],
                "payments": [
                    {
                        "payment_type": p.payment_type,
                        "amount": float(p.amount),
                        "credit_installments": p.credit_installments,
                        "boletos": [
                            {
                                "value": float(b.value),
                                "due_date": str(b.due_date)
                            }
                            for b in boletos if b.sale_id == sale.id
                        ] if p.payment_type == "BO" else []
                    }
                    for p in payments
                ]
            }
        })

    except Sale.DoesNotExist:
        return JsonResponse({"success": False, "error": "Venda não encontrada."}, status=404)

def get_product_prices(request, product_id):
    try:
        # Últimos 5 preços de venda
        sales = list(
            SaleProduct.objects.filter(product_id=product_id)
            .order_by('-id')[:5]
            .values('price', 'quantity', 'sale__created')
        )

        # Últimos 5 preços de compra
        purchases = list(
            PurchaseProduct.objects.filter(product_id=product_id)
            .order_by('-id')[:5]
            .values('price', 'quantity', 'purchase__created')
        )

        return JsonResponse({
            "sales": sales,
            "purchases": purchases
        })
    except Product.DoesNotExist:
        return JsonResponse({"error": "Produto não encontrado"}, status=404)

@login_required
def get_product_by_id(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    product_data = {
        'id': product.id,
        'name': product.name,
        'weight': product.weight,
        'unit_type': product.unit_type,
        'size': product.size,
        'quantity_per_package': product.quantity_per_package,
        'observations': product.observations,
        'manufacturer_id': product.manufacturer.id if product.manufacturer else None,
        'manufacturer_name': product.manufacturer.name if product.manufacturer else None,
        'category_id': product.category.id if product.category else None,
        'category_name': product.category.name if product.category else None,
    }

    return JsonResponse(product_data)